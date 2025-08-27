import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import os

# Load the existing workbook
wb = load_workbook("Graston_Technique_Dashboard.xlsx")

# Get the dataframes from the original CSV files for reference
table_data_path = "Table Data"
dataframes = {}

# Load dataframes for reference (just the headers to understand structure)
csv_files = [
    "addresses.csv",
    "attachments.csv",
    "businesses.csv",
    "event_registrations.csv",
    "events.csv",
    "instrument_sets.csv",
    "licenses.csv",
    "modules.csv",
    "orders.csv",
    "shipments.csv",
    "users.csv",
    "venues.csv",
    "vouchers.csv"
]

for csv_file in csv_files:
    try:
        file_path = os.path.join(table_data_path, csv_file)
        df = pd.read_csv(file_path, nrows=5)  # Just read first 5 rows to get structure
        sheet_name = csv_file.replace(".csv", "")
        dataframes[sheet_name] = df
        print(f"Loaded structure for {csv_file}")
    except Exception as e:
        print(f"Error loading {csv_file}: {str(e)}")

# Get the User Dashboard sheet
dashboard_sheet = wb["User Dashboard"]

# Enhance the User Dashboard with lookup formulas
# Add a cell to store the selected user ID
dashboard_sheet["A30"] = "Selected User ID:"
dashboard_sheet["B30"] = "=INDEX(users[ID], MATCH(B3, users[first_name]&\" \"&users[last_name], 0))"

# Add user details section with XLOOKUP formulas
dashboard_sheet["A6"] = "Full Name"
dashboard_sheet["B6"] = "Email"
dashboard_sheet["C6"] = "Phone"
dashboard_sheet["D6"] = "Type"
dashboard_sheet["E6"] = "Last Login"

# Add formulas to get user details (we'll use INDEX/MATCH since XLOOKUP might not be available)
dashboard_sheet["A7"] = "=IF(B3<>\"Select a user\", B3, \"\")"
dashboard_sheet["B7"] = "=IF(B3<>\"Select a user\", INDEX(users[email], MATCH(B3, users[first_name]&\" \"&users[last_name], 0)), \"\")"
dashboard_sheet["C7"] = "=IF(B3<>\"Select a user\", INDEX(users[phone_number], MATCH(B3, users[first_name]&\" \"&users[last_name], 0)), \"\")"
dashboard_sheet["D7"] = "=IF(B3<>\"Select a user\", INDEX(users[type], MATCH(B3, users[first_name]&\" \"&users[last_name], 0)), \"\")"
dashboard_sheet["E7"] = "=IF(B3<>\"Select a user\", INDEX(users[last_login], MATCH(B3, users[first_name]&\" \"&users[last_name], 0)), \"\")"

# Add licenses section with FILTER-like functionality
dashboard_sheet["A11"] = "License Number"
dashboard_sheet["B11"] = "Type"
dashboard_sheet["C11"] = "State"
dashboard_sheet["D11"] = "Expiration Date"
dashboard_sheet["E11"] = "Status"

# Add formulas for licenses (using INDEX/AGGREGATE for FILTER-like behavior)
dashboard_sheet["A12"] = "=IF(B3<>\"Select a user\", INDEX(licenses[number], AGGREGATE(15, 6, (ROW(licenses[user_id])-ROW(licenses[#Headers]))/(licenses[user_id]=B30), 1)), \"\")"
dashboard_sheet["B12"] = "=IF(B3<>\"Select a user\", INDEX(licenses[type], AGGREGATE(15, 6, (ROW(licenses[user_id])-ROW(licenses[#Headers]))/(licenses[user_id]=B30), 1)), \"\")"
dashboard_sheet["C12"] = "=IF(B3<>\"Select a user\", INDEX(licenses[state], AGGREGATE(15, 6, (ROW(licenses[user_id])-ROW(licenses[#Headers]))/(licenses[user_id]=B30), 1)), \"\")"
dashboard_sheet["D12"] = "=IF(B3<>\"Select a user\", INDEX(licenses[expiration], AGGREGATE(15, 6, (ROW(licenses[user_id])-ROW(licenses[#Headers]))/(licenses[user_id]=B30), 1)), \"\")"

# Add orders section
dashboard_sheet["A16"] = "Order ID"
dashboard_sheet["B16"] = "Status"
dashboard_sheet["C16"] = "Date"
dashboard_sheet["D16"] = "Total Price"

# Add formulas for orders
dashboard_sheet["A17"] = "=IF(B3<>\"Select a user\", INDEX(orders[id], AGGREGATE(15, 6, (ROW(orders[user_id])-ROW(orders[#Headers]))/(orders[user_id]=B30), 1)), \"\")"
dashboard_sheet["B17"] = "=IF(B3<>\"Select a user\", INDEX(orders[status], AGGREGATE(15, 6, (ROW(orders[user_id])-ROW(orders[#Headers]))/(orders[user_id]=B30), 1)), \"\")"
dashboard_sheet["C17"] = "=IF(B3<>\"Select a user\", INDEX(orders[created_at], AGGREGATE(15, 6, (ROW(orders[user_id])-ROW(orders[#Headers]))/(orders[user_id]=B30), 1)), \"\")"
dashboard_sheet["D17"] = "=IF(B3<>\"Select a user\", INDEX(orders[total_price], AGGREGATE(15, 6, (ROW(orders[user_id])-ROW(orders[#Headers]))/(orders[user_id]=B30), 1)), \"\")"

# Add event registrations section
dashboard_sheet["A21"] = "Event Name"
dashboard_sheet["B21"] = "Status"
dashboard_sheet["C21"] = "Date"

# Add formulas for event registrations
dashboard_sheet["A22"] = "=IF(B3<>\"Select a user\", INDEX(events[name], AGGREGATE(15, 6, (ROW(event_registrations[user_id])-ROW(event_registrations[#Headers]))/(event_registrations[user_id]=B30), 1)), \"\")"
dashboard_sheet["B22"] = "=IF(B3<>\"Select a user\", INDEX(event_registrations[status], AGGREGATE(15, 6, (ROW(event_registrations[user_id])-ROW(event_registrations[#Headers]))/(event_registrations[user_id]=B30), 1)), \"\")"
dashboard_sheet["C22"] = "=IF(B3<>\"Select a user\", INDEX(events[first_day_start], AGGREGATE(15, 6, (ROW(event_registrations[user_id])-ROW(event_registrations[#Headers]))/(event_registrations[user_id]=B30), 1)), \"\")"

# Add conditional formatting for licenses expiration
# Note: In a real implementation, we would add actual conditional formatting rules
# For now, we'll add a note about where this would be implemented
dashboard_sheet["F11"] = "Note: Conditional formatting would highlight expired licenses in red"

# Add slicers information
dashboard_sheet["A25"] = "Slicers:"
dashboard_sheet["A26"] = "For filtering data by different criteria, slicers can be added to the dashboard."
dashboard_sheet["A27"] = "In Excel, go to Insert > Slicer and select the tables you want to filter."

# Create team-specific views
# Customer Support view
customer_support_sheet = wb.create_sheet(title="Customer Support View")
customer_support_sheet["A1"] = "Customer Support Dashboard"
customer_support_sheet["A1"].font = Font(size=16, bold=True)
customer_support_sheet.merge_cells("A1:E1")

customer_support_sheet["A3"] = "User Details"
customer_support_sheet["A3"].font = Font(bold=True, size=14)
customer_support_sheet["A4"] = "Full Name"
customer_support_sheet["B4"] = "Email"
customer_support_sheet["C4"] = "Phone"
customer_support_sheet["D4"] = "Licenses"
customer_support_sheet["E4"] = "Expiration"

customer_support_sheet["A7"] = "Orders"
customer_support_sheet["A7"].font = Font(bold=True, size=14)
customer_support_sheet["A8"] = "Order ID"
customer_support_sheet["B8"] = "Status"
customer_support_sheet["C8"] = "Date"
customer_support_sheet["D8"] = "Total Price"

customer_support_sheet["A11"] = "Shipments"
customer_support_sheet["A11"].font = Font(bold=True, size=14)
customer_support_sheet["A12"] = "Order ID"
customer_support_sheet["B12"] = "Tracking Number"
customer_support_sheet["C12"] = "Shipment Date"

# Sales view
sales_sheet = wb.create_sheet(title="Sales View")
sales_sheet["A1"] = "Sales Dashboard"
sales_sheet["A1"].font = Font(size=16, bold=True)
sales_sheet.merge_cells("A1:E1")

sales_sheet["A3"] = "User Details"
sales_sheet["A3"].font = Font(bold=True, size=14)
sales_sheet["A4"] = "Name"
sales_sheet["B4"] = "Email"
sales_sheet["C4"] = "Phone"
sales_sheet["D4"] = "Location"
sales_sheet["E4"] = "Preferred Provider"

sales_sheet["A7"] = "Order History"
sales_sheet["A7"].font = Font(bold=True, size=14)
sales_sheet["A8"] = "Order ID"
sales_sheet["B8"] = "Date"
sales_sheet["C8"] = "Total Price"
sales_sheet["D8"] = "Items"

sales_sheet["A11"] = "Event Registrations"
sales_sheet["A11"].font = Font(bold=True, size=14)
sales_sheet["A12"] = "Event Name"
sales_sheet["B12"] = "Date"
sales_sheet["C12"] = "Status"

# Marketing view
marketing_sheet = wb.create_sheet(title="Marketing View")
marketing_sheet["A1"] = "Marketing Dashboard"
marketing_sheet["A1"].font = Font(size=16, bold=True)
marketing_sheet.merge_cells("A1:E1")

marketing_sheet["A3"] = "User Demographics"
marketing_sheet["A3"].font = Font(bold=True, size=14)
marketing_sheet["A4"] = "User Type"
marketing_sheet["B4"] = "Location"
marketing_sheet["C4"] = "Business"

marketing_sheet["A7"] = "Event History"
marketing_sheet["A7"].font = Font(bold=True, size=14)
marketing_sheet["A8"] = "Event Type"
marketing_sheet["B8"] = "Registration Date"
marketing_sheet["C8"] = "Status"

marketing_sheet["A11"] = "Vouchers"
marketing_sheet["A11"].font = Font(bold=True, size=14)
marketing_sheet["A12"] = "Voucher ID"
marketing_sheet["B12"] = "Status"
marketing_sheet["C12"] = "Expiration"

# Save the enhanced workbook
wb.save("Graston_Technique_Dashboard_Enhanced.xlsx")
print("Dashboard enhanced successfully!")