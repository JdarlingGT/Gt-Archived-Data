import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Define the path to the Table Data folder
table_data_path = "Table Data"

# List of CSV files to process
csv_files = [
    "addresses.csv",
    "attachments.csv",
    "businesses.csv",
    "event_registrations.csv",
    "events.csv",
    "instrument_sets.csv",
    "licenses.csv",
    "modules.csv",
    "orders.csv",
    "shipments.csv",
    "users.csv",
    "venues.csv",
    "vouchers.csv"
]

# Create a new Excel workbook
wb = Workbook()

# Remove the default sheet
wb.remove(wb.active)

# Function to create a table from CSV data
def create_table_from_csv(filename, sheet_name=None):
    # If no sheet name is provided, use the filename without extension
    if sheet_name is None:
        sheet_name = filename.replace(".csv", "")

    print(f"Processing {filename}...")

    # Read the CSV file with error handling
    try:
        file_path = os.path.join(table_data_path, filename)
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252']
        df = None

        for encoding in encodings:
            try:
                # Read only first 1000 rows for testing
                df = pd.read_csv(file_path, encoding=encoding, nrows=1000)
                print(f"Successfully read {filename} with {encoding} encoding")
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"Error reading {filename} with {encoding}: {str(e)}")
                continue

        if df is None:
            # If all encodings fail, try with error handling
            df = pd.read_csv(file_path, encoding='utf-8', encoding_errors='ignore', nrows=1000)
            print(f"Read {filename} with error handling")

    except Exception as e:
        print(f"Error reading {filename}: {str(e)}")
        return None, None

    # Create a new worksheet
    try:
        ws = wb.create_sheet(title=sheet_name)
    except Exception as e:
        # If sheet name is too long, truncate it
        sheet_name = sheet_name[:31]
        ws = wb.create_sheet(title=sheet_name)

    # Write the dataframe to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Create a table
    if len(df) > 0:
        # Define the table range
        table_range = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"

        # Create the table
        tab = Table(displayName=sheet_name.replace(" ", "_").replace("-", "_"), ref=table_range)

        # Add a default table style
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style

        # Add the table to the worksheet
        ws.add_table(tab)

    return ws, df

# Create tables for all CSV files
tables = {}
dataframes = {}

for csv_file in csv_files:
    try:
        sheet_name = csv_file.replace(".csv", "")
        ws, df = create_table_from_csv(csv_file, sheet_name)
        if ws is not None and df is not None:
            tables[sheet_name] = ws
            dataframes[sheet_name] = df
            print(f"Created table for {csv_file}")
        else:
            print(f"Failed to create table for {csv_file}")
    except Exception as e:
        print(f"Error processing {csv_file}: {str(e)}")

# Create the User Dashboard sheet
dashboard_sheet = wb.create_sheet(title="User Dashboard", index=0)

# Add a title
dashboard_sheet["A1"] = "Graston Technique Clinician Dashboard"
dashboard_sheet["A1"].font = Font(size=16, bold=True)
dashboard_sheet["A1"].alignment = Alignment(horizontal="center")
dashboard_sheet.merge_cells("A1:E1")

# Add user selection dropdown
dashboard_sheet["A3"] = "Select User:"
dashboard_sheet["A3"].font = Font(bold=True)

# Create a data validation dropdown for users
# Get the list of users for the dropdown
if "users" in dataframes and dataframes["users"] is not None and not dataframes["users"].empty:
    users_df = dataframes["users"]
    # Create a list of user names for the dropdown
    user_names = []
    for index, row in users_df.iterrows():
        first_name = row["first_name"] if pd.notna(row["first_name"]) else ""
        last_name = row["last_name"] if pd.notna(row["last_name"]) else ""
        email = row["email"] if pd.notna(row["email"]) else ""
        user_name = f"{first_name} {last_name}".strip()
        if not user_name and email:
            user_name = email
        if user_name:
            user_names.append(user_name)

    # Add the user names to a hidden sheet for the dropdown
    if "User_List" not in wb.sheetnames:
        user_list_sheet = wb.create_sheet(title="User_List")
        for i, name in enumerate(user_names, 1):
            user_list_sheet[f"A{i}"] = name

    # Create data validation for the dropdown
    if len(user_names) > 0:
        dv = DataValidation(type="list", formula1=f"User_List!$A$1:$A${len(user_names)}", allow_blank=True)
        dashboard_sheet.add_data_validation(dv)
        dv.add("B3")

    # Add a label for the selected user
    dashboard_sheet["B3"] = "Select a user"

# Add sections for different data views
dashboard_sheet["A5"] = "User Details"
dashboard_sheet["A5"].font = Font(bold=True, size=14)

dashboard_sheet["A10"] = "Licenses"
dashboard_sheet["A10"].font = Font(bold=True, size=14)

dashboard_sheet["A15"] = "Orders"
dashboard_sheet["A15"].font = Font(bold=True, size=14)

dashboard_sheet["A20"] = "Event Registrations"
dashboard_sheet["A20"].font = Font(bold=True, size=14)

# Add headers for user details
dashboard_sheet["A6"] = "Full Name"
dashboard_sheet["B6"] = "Email"
dashboard_sheet["C6"] = "Phone"
dashboard_sheet["D6"] = "Type"
dashboard_sheet["E6"] = "Last Login"

# Add headers for licenses
dashboard_sheet["A11"] = "License Number"
dashboard_sheet["B11"] = "Type"
dashboard_sheet["C11"] = "State"
dashboard_sheet["D11"] = "Expiration Date"
dashboard_sheet["E11"] = "Status"

# Add headers for orders
dashboard_sheet["A16"] = "Order ID"
dashboard_sheet["B16"] = "Status"
dashboard_sheet["C16"] = "Date"
dashboard_sheet["D16"] = "Total Price"

# Add headers for event registrations
dashboard_sheet["A21"] = "Event Name"
dashboard_sheet["B21"] = "Status"
dashboard_sheet["C21"] = "Date"

# Save the workbook
try:
    wb.save("Graston_Technique_Dashboard.xlsx")
    print("Dashboard created successfully!")
except Exception as e:
    print(f"Error saving dashboard: {str(e)}")